# ==========================================================
# main.py — CNC Energy, Cost & Sustainability Model
# ==========================================================

import math
import numpy as np
import pandas as pd
from pathlib import Path
import tensorflow as tf
import joblib
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, Series



# ==========================================================
# COSTANTI
# ==========================================================
ACC_LINEAR = 100

COSTO_ORARIO = 60.0       # €/h
COSTO_KWH = 0.40          # €/kWh
COSTO_UTENSILE = 30.0     # €/utensile

# ==========================================================
# PATH
# ==========================================================
BASE_DIR = Path(__file__).resolve().parent
MODELS_DIR = BASE_DIR / "models"

# ==========================================================
# LOAD MODELS & SCALERS
# ==========================================================
def load_models_and_scalers():
    models = {
        "G0": tf.keras.models.load_model(MODELS_DIR / "model_G0.h5", compile=False),
        "G1": tf.keras.models.load_model(MODELS_DIR / "model_G1.h5", compile=False),
        "G2": tf.keras.models.load_model(MODELS_DIR / "model_G2.h5", compile=False),
    }

    scalers = {
        "G0": (
            joblib.load(MODELS_DIR / "scalerX2_G0.pkl"),
            joblib.load(MODELS_DIR / "scalerY2_G0.pkl"),
        ),
        "G1": (
            joblib.load(MODELS_DIR / "scalerX2_G1.pkl"),
            joblib.load(MODELS_DIR / "scalerY2_G1.pkl"),
        ),
        "G2": (
            joblib.load(MODELS_DIR / "scalerX2_G2.pkl"),
            joblib.load(MODELS_DIR / "scalerY2_G2.pkl"),
        ),
    }

    return models, scalers

# ==========================================================
# PARSER
# ==========================================================
from parser_module import parse_lavorazioni_regex

# ==========================================================
# OVERRIDE / TIME / MRR
# ==========================================================
from override_module import apply_override, recompute_time, recompute_MRR

# ==========================================================
# PREDICT POWER
# ==========================================================
def predict_power(df, models, scalers):
    preds = []

    for _, r in df.iterrows():

        if r["movimento"] == "G0":
            X = np.array([[r["xv"], r["yv"], r["zv"], r["S"]]])
            Xs = scalers["G0"][0].transform(X)
            ys = models["G0"].predict(Xs, verbose=0)
            y = scalers["G0"][1].inverse_transform(ys)[0, 0]

        elif r["movimento"] == "G1":
            X = np.array([[r["xv"], r["yv"], r["zv"],
                           r["F"], r["S"], r["ideal_MRR"]]])
            Xs = scalers["G1"][0].transform(X)
            ys = models["G1"].predict(Xs, verbose=0)
            y = scalers["G1"][1].inverse_transform(ys)[0, 0]

        elif r["movimento"] == "G2":
            X = np.array([[r["R"], r["F"], r["S"], r["ideal_MRR"]]])
            Xs = scalers["G2"][0].transform(X)
            ys = models["G2"].predict(Xs, verbose=0)
            y = scalers["G2"][1].inverse_transform(ys)[0, 0]

        else:
            y = 0.0

        preds.append(float(y))

    df = df.copy()
    df["Power_pred"] = preds
    return df


# ==========================================================
# PIPELINE COMPLETA
# ==========================================================
def simulate(df_base: pd.DataFrame, k: float, models, scalers) -> pd.DataFrame:
    df = apply_override(df_base, k)
    df = recompute_time(df)
    df = recompute_MRR(df)
    df = predict_power(df, models, scalers)
    return df

# ==========================================================
# KPI
# ==========================================================
def compute_kpi(df, k, costo_orario, costo_kwh, costo_utensile):

    import math
    import pandas as pd

    # ===== TEMPO & ENERGIA =====
    T_tot = float(df["t_ideal_cumul"].max() or 0.0)          # [s]
    E_tot = float((df["Power_pred"] * df["ideal_seg_time"]).sum() or 0.0)  # [J]
    P_mean = E_tot / T_tot if T_tot > 0 else 0.0             # [W]

    # ===== COSTI =====
    T_tot_h = T_tot / 3600.0
    C_macchina = T_tot_h * costo_orario

    E_kWh = E_tot / 3.6e6
    C_energia = E_kWh * costo_kwh

    # ===== COSTO UTENSILE (Taylor) =====
    K_TAYLOR = 0.2
    C_TAYLOR = 1400.0

    D_max = df["diametro_utensile"].dropna().max()
    if pd.isna(D_max) or D_max <= 0:
        D_max = 1.0

    vita_consumata = 0.0
    for _, r in df[df["movimento"].isin(["G1", "G2"])].iterrows():
        S = r["S"]
        t = r["ideal_seg_time"]
        if S > 0 and t > 0:
            v_t = (S * math.pi * D_max) / 1000.0 * k
            if v_t > 0:
                T_ut = (C_TAYLOR / v_t) ** (1 / K_TAYLOR) * 60.0
                vita_consumata += t / T_ut

    C_utensile = vita_consumata * costo_utensile
    C_totale = C_macchina + C_energia + C_utensile

    # ===== EMISSIONI =====
    EF_ELETTRICITA = 0.25
    Em_macchina = E_kWh * EF_ELETTRICITA

    densita_tungsteno = 19e-6
    L_ut = 100.0
    fattore_geom = 0.7
    EF_tungsteno = 50.0

    volume_ut = (math.pi / 4) * D_max**2 * L_ut * fattore_geom
    massa_ut = volume_ut * densita_tungsteno
    Em_utensile = massa_ut * vita_consumata * EF_tungsteno

    Em_tot = Em_macchina + Em_utensile

    return {
    "k": k,
    "tempo_s": T_tot,

    # Energia
    "energia_J": E_tot,
    "potenza_media_W": P_mean,

    # Costi
    "costo_macchina_€": C_macchina,
    "costo_energia_€": C_energia,
    "costo_utensile_€": C_utensile,
    "costo_totale_€": C_totale,

    # Emissioni
    "emissioni_macchina_kgCO2e": Em_macchina,
    "emissioni_utensile_kgCO2e": Em_utensile,
    "emissioni_totali_kgCO2e": Em_tot,
}


# ==========================================================
# SWEEP k
# ==========================================================
def sweep_k(df_base, k_values, models, scalers):
    results = []

    for k in k_values:
        df = apply_override(df_base, k)
        df = recompute_time(df)
        df = recompute_MRR(df)
        df = predict_power(df, models, scalers)

        kpi = compute_kpi(
            df, k,
            COSTO_ORARIO,
            COSTO_KWH,
            COSTO_UTENSILE
        )
        results.append(kpi)

    return pd.DataFrame(results)

def find_optimal_k(df_kpi):
    return {
        "tempo": float(df_kpi.loc[df_kpi["tempo_s"].idxmin(), "k"]),
        "costo": float(df_kpi.loc[df_kpi["costo_totale_€"].idxmin(), "k"]),
        "emissioni": float(df_kpi.loc[df_kpi["emissioni_totali_kgCO2e"].idxmin(), "k"]),
    }

# ==========================================================
# DESIRABILITY FUNCTION (Derringer–Suich)
# ==========================================================
def desirability_min(y, y_min, y_max, s=1.0):
    if y <= y_min:
        return 1.0
    if y >= y_max:
        return 0.0
    return ((y_max - y) / (y_max - y_min)) ** s


def compute_desirability(df_kpi, weights, shapes):

    df = df_kpi.copy()

    limits = {
        "tempo": (df["tempo_s"].min(), df["tempo_s"].max()),
        "costo": (df["costo_totale_€"].min(), df["costo_totale_€"].max()),
        "emissioni": (
            df["emissioni_totali_kgCO2e"].min(),
            df["emissioni_totali_kgCO2e"].max(),
        ),
    }

    D = []

    for _, r in df.iterrows():

        d_t = desirability_min(
            r["tempo_s"], *limits["tempo"], shapes["tempo"]
        )
        d_c = desirability_min(
            r["costo_totale_€"], *limits["costo"], shapes["costo"]
        )
        d_e = desirability_min(
            r["emissioni_totali_kgCO2e"], *limits["emissioni"], shapes["emissioni"]
        )

        D_global = (
            (d_t ** weights["tempo"])
            * (d_c ** weights["costo"])
            * (d_e ** weights["emissioni"])
        ) ** (1 / sum(weights.values()))

        D.append(D_global)

    df["Desirability"] = D
    return df


def find_optimal_k_desirability(df):
    idx = df["Desirability"].idxmax()
    return df.loc[idx, "k"]


from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

# ==========================================================
# EXPORT
# ==========================================================


def export_sweep_to_excel(df_kpi, df_des, k_opt, filename):

    # ================= SCRITTURA DATI =================
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:

        df_kpi.to_excel(writer, sheet_name="KPI", index=False)

        df_kpi[["k", "tempo_s"]].to_excel(
            writer, sheet_name="Tempo", index=False
        )

        df_kpi[
            ["k",
             "costo_macchina_€",
             "costo_energia_€",
             "costo_utensile_€",
             "costo_totale_€"]
        ].to_excel(writer, sheet_name="Costi", index=False)

        df_kpi[
            ["k",
             "emissioni_macchina_kgCO2e",
             "emissioni_utensile_kgCO2e",
             "emissioni_totali_kgCO2e"]
        ].to_excel(writer, sheet_name="Emissioni", index=False)

        df_des[["k", "Desirability"]].to_excel(
            writer, sheet_name="Desirability", index=False
        )

    # ================= GRAFICO DESIRABILITY =================
    wb = load_workbook(filename)
    ws = wb["Desirability"]

    chart = LineChart()
    chart.title = "Desirability globale vs k"
    chart.y_axis.title = "Desirability"
    chart.x_axis.title = "k"

    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # ===== LINEA k OTTIMO =====
    ws["C1"] = "k_opt"
    for i in range(2, ws.max_row + 1):
        ws[f"C{i}"] = (
            ws[f"B{i}"].value if ws[f"A{i}"].value == k_opt else None
        )

    opt_data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
    chart.add_data(opt_data, titles_from_data=True)

    ws.add_chart(chart, "E2")

    wb.save(filename)


if __name__ == "__main__":

    print("Loading models and scalers...")
    models, scalers = load_models_and_scalers()

    print("Parsing G-code...")
    df_base = parse_lavorazioni_regex("Test60x40_11.cnc")

    print("Running sweep k...")
    k_values = [round(k, 2) for k in np.arange(0.5, 1.51, 0.1)]
    df_kpi = sweep_k(df_base, k_values, models, scalers)

    # ===== DESIRABILITY SETUP =====
    weights = {
        "tempo": 1,
        "costo": 1,
        "emissioni": 2
    }

    shapes = {
        "tempo": 1.0,
        "costo": 1.0,
        "emissioni": 1.5
    }

    df_des = compute_desirability(df_kpi, weights, shapes)
    k_opt = find_optimal_k_desirability(df_des)

    print(f"k ottimo multi-obiettivo = {k_opt}")

    export_sweep_to_excel(
        df_kpi,
        df_des,
        k_opt,
        "sweep_k_results.xlsx"
    )

    print("Excel multi-sheet con Desirability generato.")


    print("Loading models and scalers...")
    models, scalers = load_models_and_scalers()

    print("Parsing G-code...")
    df_base = parse_lavorazioni_regex("Test60x40_11.cnc")

    print("Running sweep k...")
    k_values = [round(k, 2) for k in np.arange(0.5, 1.51, 0.1)]
    df_kpi = sweep_k(df_base, k_values, models, scalers)

    print(df_kpi)

    k_opt = find_optimal_k(df_kpi)
    print("k ottimi:", k_opt)

    export_sweep_to_excel(df_kpi, k_opt, "sweep_k_results.xlsx")
    print("Excel multi-sheet generated.")
